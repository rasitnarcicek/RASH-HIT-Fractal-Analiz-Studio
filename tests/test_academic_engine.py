# SPDX-License-Identifier: Apache-2.0
# Copyright 2026 Mehmet Raşit Narçiçek
"""
test_academic_engine.py - Test suite using strictly input_svgs/ folder SVG files.
"""
from __future__ import annotations

import json
from pathlib import Path
import pytest

from backend.processor import AnalysisProcessor
from backend.batch_processor import run_batch_analysis
from backend.svg_health import inspect_svg_health
from backend.regression import compute_loglog_regression

INPUT_SVGS_DIR = Path("input_svgs").resolve()


def test_svg_health_inspection_valid():
    sample_file = INPUT_SVGS_DIR / "16A.svg"
    res = inspect_svg_health(sample_file)
    assert res.is_valid_xml is True
    assert res.total_shape_elements > 0
    assert res.suitability in ("High", "Moderate")


def test_svg_health_reports_health_ms():
    """RASH-HIT Fractal Engine (Realtime Metrics): inspect_svg_health must time itself and
    expose health_ms on the result and its dict payload."""
    sample_file = INPUT_SVGS_DIR / "16A.svg"
    res = inspect_svg_health(sample_file)
    assert res.health_ms >= 0.0
    payload = res.to_dict()
    assert "health_ms" in payload
    assert payload["health_ms"] >= 0.0

    # Failed inspections (missing file) must also report a measured duration.
    missing = inspect_svg_health(INPUT_SVGS_DIR / "does_not_exist.svg")
    assert missing.health_ms >= 0.0
    assert missing.errors


def test_svg_health_inspection_empty(tmp_path):
    empty_file = tmp_path / "empty.svg"
    empty_file.write_bytes(b"")
    res = inspect_svg_health(empty_file)
    assert res.is_empty is True
    assert res.total_shape_elements == 0
    assert res.suitability == "Unsuitable"
    assert len(res.errors) > 0


def test_svg_health_inspection_broken(tmp_path):
    broken_file = tmp_path / "broken.svg"
    broken_file.write_text("<svg><path d='M 10 10'</svg>", encoding="utf-8")
    res = inspect_svg_health(broken_file)
    assert res.is_valid_xml is False
    assert res.suitability == "Unsuitable"
    assert len(res.errors) > 0


def test_invalid_file_path_handling(tmp_path):
    invalid_file = Path("input_svgs/non_existent_file.svg")
    proc = AnalysisProcessor(input_path=invalid_file, output_dir=tmp_path)
    exec_res = proc.run()
    assert exec_res.status == "FAILED"
    assert len(exec_res.errors) > 0
    assert "not found" in exec_res.errors[0].lower()


def test_single_file_processor_execution(tmp_path):
    sample_file = INPUT_SVGS_DIR / "16A.svg"
    proc = AnalysisProcessor(input_path=sample_file, output_dir=tmp_path, levels=3)
    res = proc.run()
    assert res.status == "SUCCESS"
    assert res.fractal_dimension > 0
    assert res.r_squared > 0
    assert (tmp_path / "16A" / "result.json").exists()


def test_academic_report_integration(tmp_path):
    sample_file = INPUT_SVGS_DIR / "16A.svg"
    proc = AnalysisProcessor(input_path=sample_file, output_dir=tmp_path, levels=3)
    res = proc.run()
    assert res.status == "SUCCESS"
    report_file = tmp_path / "16A" / "report" / "report.html"
    assert report_file.exists()
    html_content = report_file.read_text(encoding="utf-8")
    assert "header" in html_content
    assert "kpi-grid" in html_content


def test_academic_report_markdown_links(tmp_path):
    """report.md generation was removed (dead-link cleanup): the report.html
    must no longer reference a report.md file that is never produced."""
    sample_file = INPUT_SVGS_DIR / "16A.svg"
    proc = AnalysisProcessor(input_path=sample_file, output_dir=tmp_path, levels=3)
    res = proc.run()
    report_file = tmp_path / "16A" / "report" / "report.html"
    assert report_file.exists()
    content = report_file.read_text(encoding="utf-8")
    assert "report.md" not in content, "Dead report.md link must not be emitted"
    assert not (tmp_path / "16A" / "report" / "report.md").exists()


def test_academic_report_theme_parity(tmp_path):
    sample_file = INPUT_SVGS_DIR / "16A.svg"
    proc = AnalysisProcessor(input_path=sample_file, output_dir=tmp_path, levels=3)
    res = proc.run()
    report_file = tmp_path / "16A" / "report" / "report.html"
    content = report_file.read_text(encoding="utf-8")
    assert "--bg" in content
    assert "data-theme" in content


def test_academic_report_with_manifest(tmp_path):
    sample_file = INPUT_SVGS_DIR / "16A.svg"
    proc = AnalysisProcessor(input_path=sample_file, output_dir=tmp_path, levels=3)
    res = proc.run()
    manifest_file = tmp_path / "16A" / "manifest" / "manifest.json"
    assert manifest_file.exists()


def test_pdf_report_export_mocked(tmp_path):
    sample_file = INPUT_SVGS_DIR / "16A.svg"
    proc = AnalysisProcessor(input_path=sample_file, output_dir=tmp_path, levels=3)
    res = proc.run()
    report_dir = tmp_path / "16A" / "report"
    assert report_dir.exists()


def test_tables_viewer_not_generated(tmp_path):
    """tables/tables.html and tables/tables_data.json are no longer produced
    (user request 2026-08-05): the workbook is the single spreadsheet artifact."""
    sample_file = INPUT_SVGS_DIR / "16A.svg"
    proc = AnalysisProcessor(input_path=sample_file, output_dir=tmp_path, levels=3)
    res = proc.run()
    assert not (tmp_path / "16A" / "tables" / "tables.html").exists()
    assert not (tmp_path / "16A" / "tables" / "tables_data.json").exists()
    # No per-level cells.xlsx tables either.
    tables_dir = tmp_path / "16A" / "tables"
    if tables_dir.exists():
        assert not list(tables_dir.glob("*_cells.xlsx"))
    # The workbook must still be produced.
    assert (tmp_path / "16A" / "excel" / "workbook.xlsx").exists()


def test_tables_viewer_has_manifest(tmp_path):
    sample_file = INPUT_SVGS_DIR / "16A.svg"
    proc = AnalysisProcessor(input_path=sample_file, output_dir=tmp_path, levels=3)
    res = proc.run()
    manifest_file = tmp_path / "16A" / "manifest" / "manifest.json"
    assert manifest_file.exists()


def test_batch_processing_execution(tmp_path):
    # Use levels=1 for ultra-fast execution during batch tests
    batch_res = run_batch_analysis(folder_path=INPUT_SVGS_DIR, output_dir=tmp_path, levels=1)

    assert batch_res.total_files == len(list(INPUT_SVGS_DIR.glob("*.svg")))
    assert batch_res.successful_count > 0
    assert not (Path(batch_res.output_dir) / "batch_summary.json").exists()
    assert not (Path(batch_res.output_dir) / "batch_summary.csv").exists()
    assert not (Path(batch_res.output_dir) / "batch_report.html").exists()
    assert (Path(batch_res.output_dir) / "individual_results").exists()


GOLDEN_DB_RANGES = {
    "16.svg": (1.50, 2.00),
    "16A.svg": (1.50, 2.00),
    "16B.svg": (1.50, 2.00),
    "16C.svg": (1.50, 2.00),
}


@pytest.mark.parametrize(
    "sample_name,min_db,max_db",
    [(name, lo, hi) for name, (lo, hi) in GOLDEN_DB_RANGES.items()],
)
def test_golden_sample_db_range(tmp_path, sample_name, min_db, max_db):
    sample_file = INPUT_SVGS_DIR / sample_name
    proc = AnalysisProcessor(input_path=sample_file, output_dir=tmp_path / sample_name, levels=3)
    res = proc.run()
    assert res.status == "SUCCESS", f"{sample_name}: {res.errors}"
    assert min_db <= res.fractal_dimension <= max_db, (
        f"{sample_name} Db={res.fractal_dimension:.4f} outside expected "
        f"[{min_db}, {max_db}]"
    )


def test_deterministic_repeated_output(tmp_path):
    sample_file = INPUT_SVGS_DIR / "16A.svg"
    r1 = AnalysisProcessor(input_path=sample_file, output_dir=tmp_path / "det_a", levels=3).run()
    r2 = AnalysisProcessor(input_path=sample_file, output_dir=tmp_path / "det_b", levels=3).run()
    assert r1.status == "SUCCESS" and r2.status == "SUCCESS"
    assert r1.fractal_dimension == pytest.approx(r2.fractal_dimension, abs=1e-9)
    assert r1.r_squared == pytest.approx(r2.r_squared, abs=1e-9)
    assert r1.computed_levels_count == r2.computed_levels_count
    assert len(r1.scale_table) == len(r2.scale_table)
    for a, b in zip(r1.scale_table, r2.scale_table):
        assert a["occupied_count"] == b["occupied_count"]
        assert a["total_count"] == b["total_count"]
        assert a["level"] == b["level"]
        assert a["grid_label"] == b["grid_label"]


def test_result_json_version_stamps(tmp_path):
    sample_file = INPUT_SVGS_DIR / "16A.svg"
    proc = AnalysisProcessor(input_path=sample_file, output_dir=tmp_path, levels=3)
    res = proc.run()
    assert res.status == "SUCCESS"
    json_data = json.loads((tmp_path / "16A" / "result.json").read_text(encoding="utf-8"))
    for key in ("engine_version", "analysis_profile_version", "generated_at", "created_at"):
        assert key in json_data, f"Missing version stamp in result.json: {key}"


def test_excel_coordinates_not_transposed(tmp_path):
    import openpyxl
    from backend.academic_exporter import (
        AnalysisReportModel,
        LevelReportModel,
        generate_excel_workbook,
    )

    lvl = LevelReportModel(
        level=1, cols=4, rows=8, grid_label="4x8",
        total_cells=32, filled_cells=3, empty_cells=29, fill_ratio=0.09375,
        occupancy_percent=9.375, cell_w=10.0, cell_h=10.0, execution_time_ms=1.0,
        filled_set={(1, 2), (2, 3), (0, 0)},  # (row, col)
    )

    model = AnalysisReportModel(
        motif="TestMotif",
        safe_name="TestMotif",
        generated_at="2026-08-10 12:00:00",
        source_file="test.svg",
        viewbox_width=40.0,
        viewbox_height=80.0,
        aspect_ratio=0.5,
        vector_geometry_count=10,
        analysis_engine="Test Engine",
        db=1.5, r2=0.99, total_time_ms=10.0,
        levels=[lvl],
    )

    out_dir = tmp_path / "excel"
    out_dir.mkdir()

    excel_file = generate_excel_workbook(model, out_dir, generate_data_maps=False)
    assert excel_file.exists()

    wb = openpyxl.load_workbook(excel_file)
    ws_map = wb["Level 01 Map"]
    ws_coord = wb["Level 01 SVG Coordinate Map"]

    # In Level 01 Map, headers are at row 6.
    # Data rows start at row 7 (r=0) up to row 14 (r=7).
    # Data columns start at column 3 (c=0) up to column 6 (c=3).
    # Checked (row, col) in filled_set:
    # (0, 0) -> Excel row 7, column 3 (C)
    # (1, 2) -> Excel row 8, column 5 (E)
    # (2, 3) -> Excel row 9, column 6 (F)

    # Assert colors on Level 01 Map
    c_0_0 = ws_map.cell(row=7, column=3)
    c_1_2 = ws_map.cell(row=8, column=5)
    c_2_3 = ws_map.cell(row=9, column=6)
    c_0_1 = ws_map.cell(row=7, column=4)  # empty

    assert c_0_0.fill.start_color.rgb in ("0060A5FA", "60A5FA")
    assert c_1_2.fill.start_color.rgb in ("0060A5FA", "60A5FA")
    assert c_2_3.fill.start_color.rgb in ("0060A5FA", "60A5FA")
    assert c_0_1.fill.start_color.rgb not in ("0060A5FA", "60A5FA")

    # Assert values on Level 01 SVG Coordinate Map
    coord_0_0 = ws_coord.cell(row=7, column=3)
    coord_1_2 = ws_coord.cell(row=8, column=5)
    coord_2_3 = ws_coord.cell(row=9, column=6)

    assert coord_0_0.value == "5.0000,5.0000"
    assert coord_1_2.value == "25.0000,15.0000"
    assert coord_2_3.value == "35.0000,25.0000"

