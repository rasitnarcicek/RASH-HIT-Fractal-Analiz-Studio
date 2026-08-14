#!/usr/bin/env python
"""Definitive orientation check for RASH-HIT workbook Level Map /
SVG Coordinate Map sheets.

Oracle = the engine's OWN filled_set (the exact data the workbook is built
from). We replay the engine on the same SVG, capture each level's filled_set,
then read back what the generated workbook actually drew and compare
cell-by-cell. This proves the fix is internally consistent (workbook == engine).

Then a second, independent oracle: rebuild a filled matrix straight from the
SVG vector geometry with shapely and compare, to confirm the engine's
filled_set itself is oriented correctly relative to the actual artwork.
"""
import sys
import re
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl
import numpy as np
from shapely.ops import unary_union

from backend.svg_loader import load_svg_geometries
from backend.intersection_hierarchical import analyze_grid_hierarchical
from backend.grid_planner import create_grid_plan

SVG = ROOT / "input_svgs" / "16.svg"
WB = ROOT / "outputs" / "verify_fix" / "16" / "excel" / "workbook.xlsx"


def read_filled_from_sheet(ws, label_re):
    """Return dict (row,col)->1 from a Map / Coordinate Map sheet by scanning
    the 'Rxx' labels in col B and 'Cxx' labels in row 6."""
    cells = {}
    max_r = max_c = -1
    for r in range(7, ws.max_row + 1):
        rl = ws.cell(row=r, column=2).value
        if not (isinstance(rl, str) and rl.startswith("R")):
            continue
        row_idx = int(rl[1:]) - 1
        for c in range(3, ws.max_column + 1):
            cl = ws.cell(row=6, column=c).value
            if not (isinstance(cl, str) and cl.startswith("C")):
                continue
            col_idx = int(cl[1:]) - 1
            # a cell is 'filled' if its (solid) fill is the filled color 60A5FA
            fg = ws.cell(row=r, column=c).fill.fgColor.rgb
            filled = (fg == "0060A5FA")  # exact match on filled blue
            cells[(row_idx, col_idx)] = 1 if filled else 0
            max_r = max(max_r, row_idx)
            max_c = max(max_c, col_idx)
    return cells, (max_r + 1, max_c + 1)


def main():
    from backend.intersection_hierarchical import compute_hierarchical_box_counting
    from backend.grid_planner import create_grid_plan
    from backend.output_profiles import load_output_profile

    raw, vw, vh = load_svg_geometries(SVG)
    shp = [g.shapely_obj for g in raw if getattr(g, "shapely_obj", None) is not None]
    geom = unary_union(shp)

    # Oracle A: the exact engine output the workbook is built from.
    grid_plan = create_grid_plan(
        svg_viewbox=(0.0, 0.0, vw, vh), svg_width=vw, svg_height=vh,
        num_levels=4, grid_mode="canvas_aspect",
    )
    profile = load_output_profile("reproducible")
    lvl_models, _ = compute_hierarchical_box_counting(
        raw, vw, vh, grid_plan=grid_plan, profile=profile,
    )

    wb = openpyxl.load_workbook(WB)

    print("=== TEST 1: workbook Level Map vs engine filled_set (oracle = engine) ===")
    t1_fail = 0
    for lvl in lvl_models:
        lvl_fmt = f"{lvl.level:02d}"
        cols, rows = lvl.cols, lvl.rows
        filled = lvl.filled_set  # set of (row, col) per source-of-truth
        expected = np.zeros((rows, cols), dtype=int)
        for (rr, cc) in filled:
            if 0 <= rr < rows and 0 <= cc < cols:
                expected[rr, cc] = 1
        ws = wb[f"Level {lvl_fmt} Map"]
        drawn_cells, shape = read_filled_from_sheet(ws, None)
        drawn = np.zeros((rows, cols), dtype=int)
        for (rr, cc), v in drawn_cells.items():
            if 0 <= rr < rows and 0 <= cc < cols:
                drawn[rr, cc] = v
        mm = int(np.sum(drawn != expected))
        status = "OK" if mm == 0 else "MISMATCH"
        if mm:
            t1_fail += 1
        print(f"  L{lvl_fmt} ({cols}x{rows}): {status} mismatches={mm}/{rows*cols}  filled_set_size={len(filled)} drawn_filled={int(drawn.sum())} expected_filled={int(expected.sum())}")

    print("\n=== TEST 2: Level Map vs SVG Coordinate Map (consistency across sheets) ===")
    t2_fail = 0
    for lvl in lvl_models:
        lvl_fmt = f"{lvl.level:02d}"
        ws_map = wb[f"Level {lvl_fmt} Map"]
        ws_coord = wb[f"Level {lvl_fmt} SVG Coordinate Map"]
        m_cells, _ = read_filled_from_sheet(ws_map, None)
        c_cells, _ = read_filled_from_sheet(ws_coord, None)
        keys = set(m_cells) | set(c_cells)
        mism = sum(1 for k in keys if m_cells.get(k, 0) != c_cells.get(k, 0))
        status = "OK" if mism == 0 else "MISMATCH"
        if mism:
            t2_fail += 1
        print(f"  L{lvl_fmt}: {status} map_vs_coord mismatches={mism}")

    print("\n=== TEST 3: engine filled_set vs independent SVG geometry (oracle = shapely) ===")
    raw, vw, vh = load_svg_geometries(SVG)
    shp = [g.shapely_obj for g in raw if getattr(g, "shapely_obj", None) is not None]
    geom = unary_union(shp)
    t3_fail = 0
    for lvl in lvl_models:
        cols, rows = lvl.cols, lvl.rows
        cw, ch = vw / cols, vh / rows
        true = np.zeros((rows, cols), dtype=int)
        for r in range(rows):
            for c in range(cols):
                box_ = __import__("shapely").geometry.box(c * cw, r * ch, (c + 1) * cw, (r + 1) * ch)
                if box_.intersects(geom):
                    true[r, c] = 1
        expected = np.zeros((rows, cols), dtype=int)
        for (rr, cc) in lvl.filled_set:
            if 0 <= rr < rows and 0 <= cc < cols:
                expected[rr, cc] = 1
        mm = int(np.sum(true != expected))
        status = "OK" if mm == 0 else "DIFF"
        if mm:
            t3_fail += 1
        print(f"  L{lvl.level:02d} ({cols}x{rows}): engine_vs_geometry {status} diffs={mm}/{rows*cols}")

    print()
    if t1_fail == 0 and t2_fail == 0 and t3_fail == 0:
        print("ALL PASS: workbook sheets correctly represent the engine's filled_set, are internally consistent, and the engine matches the artwork.")
    else:
        print(f"RESULT: T1_fail={t1_fail} T2_fail={t2_fail} T3_fail={t3_fail}")


if __name__ == "__main__":
    main()
